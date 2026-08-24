"""Deterministic, bounded parsing of CSV/XLSX business records.

This module is intentionally separate from retrieval.  It never evaluates
formulas, follows links, or performs AI-based schema inference.
"""

from __future__ import annotations

import csv
import hashlib
import json
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .models import (
    AcceptedDocument,
    DocumentType,
    SourceLocator,
    SourceLocatorKind,
    StructuredDocument,
    StructuredField,
    StructuredRecord,
    StructuredRecordType,
    StructuredSheet,
    StructuredSourceLocator,
)

PARSING_VERSION = "structured-v1"
MAX_SHEETS = 20
MAX_ROWS_PER_SHEET = 10_000
MAX_COLUMNS = 100
MAX_TOTAL_CELLS = 200_000
MAX_XLSX_UNCOMPRESSED_SIZE_BYTES = 50 * 1024 * 1024
_HEADER_ALIASES = {
    "invoice number": "invoice_number",
    "invoice #": "invoice_number",
    "inv #": "invoice_number",
    "invoice_number": "invoice_number",
    "invoice date": "invoice_date",
    "invoice_date": "invoice_date",
    "po number": "po_number",
    "po #": "po_number",
    "po_number": "po_number",
    "order date": "order_date",
    "order_date": "order_date",
    "vendor": "vendor",
    "vendor name": "vendor_name",
    "vendor_name": "vendor_name",
    "vendor id": "vendor_id",
    "vendor_id": "vendor_id",
    "item": "item",
    "product": "product_name",
    "product name": "product_name",
    "product_name": "product_name",
    "sku": "sku",
    "quantity": "quantity",
    "qty": "quantity",
    "unit price": "unit_price",
    "unit_price": "unit_price",
    "line total": "line_total",
    "line_total": "line_total",
    "price": "price",
    "description": "description",
    "unit": "unit",
    "contact": "contact",
    "terms": "terms",
}
_TOKEN = re.compile(r"[ _-]+")


class StructuredRecordError(ValueError):
    """Safe parsing error; callers should not expose source content."""


def parse_structured_document(
    document: AcceptedDocument, content: bytes
) -> StructuredDocument:
    """Parse already validated bytes into source-preserving structured records."""
    if document.document_type is DocumentType.CSV:
        tables = ((None, _parse_csv(content)),)
    elif document.document_type is DocumentType.XLSX:
        tables = _parse_xlsx(content)
    else:
        raise StructuredRecordError("Unsupported structured document type.")
    total_cells = 0
    sheets: list[StructuredSheet] = []
    for sheet_name, rows in tables:
        if not rows or not any(_cell_text(value).strip() for value in rows[0]):
            continue
        if len(rows) > MAX_ROWS_PER_SHEET + 1:
            raise StructuredRecordError("Structured record row limit exceeded.")
        headers = tuple(_header(value) for value in rows[0])
        if len(headers) > MAX_COLUMNS:
            raise StructuredRecordError("Structured record column limit exceeded.")
        if len(set(headers)) != len(headers):
            raise StructuredRecordError("Structured record headers must be unique.")
        total_cells += len(rows) * len(headers)
        if total_cells > MAX_TOTAL_CELLS:
            raise StructuredRecordError("Structured record cell limit exceeded.")
        record_type = classify_headers(headers)
        records: list[StructuredRecord] = []
        for row_number, row in enumerate(rows[1:], start=2):
            values = tuple(row[: len(headers)]) + (None,) * max(
                0, len(headers) - len(row)
            )
            if not any(_cell_text(value).strip() for value in values):
                continue
            fields = tuple(
                _field(original, canonical, value)
                for original, canonical, value in zip(
                    rows[0], headers, values, strict=True
                )
            )
            label = _record_label(record_type, fields)
            key = json.dumps(
                [(field.canonical_name, field.original_value) for field in fields],
                separators=(",", ":"),
            )
            record_id = hashlib.sha256(
                f"{PARSING_VERSION}\x1f{document.document_id}\x1f{sheet_name}\x1f{row_number}\x1f{key}".encode()
            ).hexdigest()
            records.append(
                StructuredRecord(
                    record_id,
                    document.document_id,
                    document.original_display_name,
                    document.document_type,
                    record_type,
                    row_number - 1,
                    fields,
                    StructuredSourceLocator(sheet_name, row_number, label),
                    hashlib.sha256(key.encode()).hexdigest(),
                )
            )
        sheets.append(
            StructuredSheet(
                sheet_name, tuple(_cell_text(h) for h in rows[0]), tuple(records)
            )
        )
    if not sheets or not any(sheet.records for sheet in sheets):
        raise StructuredRecordError("The structured document contains no records.")
    return StructuredDocument(
        PARSING_VERSION,
        document.document_id,
        document.original_display_name,
        document.document_type,
        document.content_hash,
        tuple(sheets),
    )


def classify_headers(headers: tuple[str, ...]) -> StructuredRecordType:
    values = set(headers)
    if {"invoice_number", "vendor"}.issubset(values) and values & {
        "item",
        "product_name",
        "line_total",
    }:
        return StructuredRecordType.INVOICE
    if {"po_number", "vendor"}.issubset(values) and values & {
        "item",
        "product_name",
        "quantity",
    }:
        return StructuredRecordType.PURCHASE_ORDER
    if "vendor_name" in values and values & {"vendor_id", "contact", "terms"}:
        return StructuredRecordType.VENDOR
    if {"sku", "product_name"}.issubset(values):
        return StructuredRecordType.PRODUCT_CATALOG
    return StructuredRecordType.GENERIC_TABULAR


def structured_evidence(document: StructuredDocument):
    """Yield record-sized pseudo-sections compatible with the chunk pipeline."""
    from .models import ExtractedDocument, ExtractedSection

    sections = []
    for sheet in document.sheets:
        for record in sheet.records:
            text = render_evidence(record)
            sections.append(
                ExtractedSection(
                    record.record_id,
                    record.document_id,
                    record.document_name,
                    record.document_type,
                    SourceLocator(
                        SourceLocatorKind.STRUCTURED_ROW,
                        sheet_name=record.source_locator.sheet_name,
                        row_number=record.source_locator.row_number,
                        record_label=record.source_locator.record_label,
                    ),
                    text,
                    record.content_hash,
                )
            )
    return ExtractedDocument(
        PARSING_VERSION,
        document.document_id,
        document.document_display_name,
        document.document_type,
        document.source_content_hash,
        tuple(sections),
    )


def render_evidence(record: StructuredRecord) -> str:
    lines = [f"Record type: {record.record_type.value.replace('_', ' ').title()}"]
    for field in record.fields:
        if field.original_value is not None:
            label = field.canonical_name.replace("_", " ").title()
            lines.append(
                f"{label}: {field.original_value}"
            )
    return "\n".join(lines)


def _parse_csv(content: bytes) -> list[list[str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise StructuredRecordError("CSV must be UTF-8 text.") from error
    if "\x00" in text:
        raise StructuredRecordError("CSV must not contain binary data.")
    try:
        dialect = csv.Sniffer().sniff(text[:8192], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    try:
        return [list(row) for row in csv.reader(text.splitlines(), dialect)]
    except csv.Error as error:
        raise StructuredRecordError("CSV is malformed.") from error


def _parse_xlsx(content: bytes) -> tuple[tuple[str, list[list[object]]], ...]:
    try:
        workbook = load_workbook(
            BytesIO(content), read_only=True, data_only=False, keep_links=False
        )
    except (InvalidFileException, OSError, ValueError, KeyError) as error:
        raise StructuredRecordError("XLSX workbook is malformed.") from error
    if len(workbook.worksheets) > MAX_SHEETS:
        raise StructuredRecordError("Structured workbook sheet limit exceeded.")
    tables = []
    for worksheet in workbook.worksheets:
        if (
            worksheet.max_row > MAX_ROWS_PER_SHEET + 1
            or worksheet.max_column > MAX_COLUMNS
        ):
            raise StructuredRecordError(
                "Structured workbook dimensions exceed V1 limits."
            )
        rows: list[list[object]] = []
        for row in worksheet.iter_rows():
            rows.append([None if cell.data_type == "f" else cell.value for cell in row])
        tables.append((worksheet.title, rows))
    return tuple(tables)


def _header(value: object) -> str:
    raw = _cell_text(value).strip()
    normalized = _TOKEN.sub(" ", raw.lower()).strip()
    return _HEADER_ALIASES.get(normalized, normalized.replace(" ", "_"))


def _field(original_header: object, canonical: str, value: object) -> StructuredField:
    original = (
        None if value is None or _cell_text(value).strip() == "" else _cell_text(value)
    )
    normalized = _normalized_value(value)
    return StructuredField(_cell_text(original_header), canonical, original, normalized)


def _normalized_value(value: object) -> str | None:
    if value is None or _cell_text(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = _cell_text(value).strip().replace(",", "").replace("$", "")
    try:
        return str(Decimal(raw))
    except InvalidOperation:
        return _cell_text(value).strip()


def _record_label(
    kind: StructuredRecordType, fields: tuple[StructuredField, ...]
) -> str | None:
    values = {field.canonical_name: field.original_value for field in fields}
    for key in ("invoice_number", "po_number", "vendor_name", "product_name", "sku"):
        if values.get(key):
            return values[key]
    return None


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    return str(value)
